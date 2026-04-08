from __future__ import annotations

import json
from pathlib import Path

from invoice_batch.config import AppConfig
from invoice_batch.domain.models import FileProcessResult, RunContext, RunSummary

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None


def _numeric(value, fallback="-"):
    """Devuelve el valor numérico real para la celda, o fallback si es None."""
    if value is None:
        return fallback
    return value


def _discount_value(value):
    """Devuelve el descuento como entero si es redondo (45), float si no (45.5), o '-'."""
    if value is None:
        return "-"
    return int(value) if value == int(value) else value


def _format_date(value) -> str:
    if value is None or value == "":
        return "-"
    return str(value)


def _is_isbn(code: str | None) -> bool:
    """Detecta si un código de producto es un ISBN-13 (13 dígitos numéricos)."""
    return bool(code and code.isdigit() and len(code) == 13)


def _apply_header_style(ws) -> None:
    """Estilo visual para la sección de cabecera (fondo azul oscuro, texto blanco)."""
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.column == 1 and cell.value:
                cell.fill = fill
                cell.font = font


def _apply_detail_header_style(ws, row_number: int) -> None:
    """Estilo para la fila de encabezado del detalle de ítems."""
    fill = PatternFill("solid", fgColor="2E75B6")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_number]:
        if cell.value:
            cell.fill = fill
            cell.font = font


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


class ExcelOutputWriter:
    def __init__(self, config: AppConfig, uploader=None) -> None:
        self.config = config
        self.uploader = uploader  # OneDriveUploader opcional

    def write_document_artifacts(self, run_context: RunContext, result: FileProcessResult) -> None:
        if self.config.processing.write_raw_json:
            self._write_raw_json(run_context, result)

        if result.document is None:
            return

        self._write_excel(run_context, result)

    def finalize_run(self, summary: RunSummary) -> None:
        output_dir = self.config.paths.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        summary_path = output_dir / f"run_summary_{summary.run_id}.json"
        summary_payload = {
            "run_id": summary.run_id,
            "started_at": summary.started_at.isoformat(),
            "finished_at": summary.finished_at.isoformat(),
            "total_files": summary.total_files,
            "success_count": summary.success_count,
            "warning_count": summary.warning_count,
            "error_count": summary.error_count,
            "skipped_count": summary.skipped_count,
            "fatal_error": summary.fatal_error,
        }
        summary_path.write_text(
            json.dumps(summary_payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )

    def _write_raw_json(self, run_context: RunContext, result: FileProcessResult) -> None:
        raw_dir = self.config.paths.output_dir / "raw_json" / run_context.run_id
        raw_dir.mkdir(parents=True, exist_ok=True)
        payload = {
            "status": result.status,
            "file_path": str(result.file_path),
            "error_message": result.error_message,
            "validation_messages": [
                {"level": msg.level, "code": msg.code, "message": msg.message}
                for msg in result.validation_messages
            ],
            "raw_payload": result.raw_payload,
        }
        (raw_dir / f"{result.file_path.stem}.json").write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )

    def _write_excel(self, run_context: RunContext, result: FileProcessResult) -> None:
        if Workbook is None:
            raise RuntimeError("openpyxl no está instalado. Ejecutar: pip install openpyxl")

        document = result.document
        fields = document.fields

        doc_type = " ".join(filter(None, [
            fields.get("document_subtype"),
            fields.get("document_letter"),
        ])) or "-"

        # Filas de cabecera — referencia visual para carga manual en Fierro
        header_rows = [
            ("Proveedor",            fields.get("vendor_name") or "-"),
            ("Tipo",                 doc_type),
            ("Nro. Comprobante",     fields.get("invoice_id") or "-"),
            ("Fecha de emisión",     _format_date(fields.get("issue_date"))),
            ("Fecha de vencimiento", _format_date(fields.get("invoice_due_date"))),
            ("Condición de pago",    fields.get("payment_terms") or "-"),
            ("Monto neto",           _numeric(fields.get("subtotal_amount"))),
            ("IVA",                  _numeric(fields.get("tax_amount"))),
            ("Monto total",          _numeric(fields.get("total_amount"))),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Comprobante"

        for label, value in header_rows:
            ws.append([label, value])

        ws.append([])  # fila en blanco separadora

        # Encabezado del detalle de ítems
        detail_header_row = ws.max_row + 1
        ws.append(["Cód. proveedor", "ISBN", "Título", "Cantidad", "Precio unitario", "Descuento %"])

        # Ítems
        for line in document.lines:
            v = line.values
            product_code = v.get("product_code") or "-"
            isbn = product_code if _is_isbn(product_code) else "-"

            ws.append([
                product_code,
                isbn,
                v.get("description"),
                v.get("quantity"),
                _numeric(v.get("unit_price")),
                _discount_value(v.get("line_discount")),
            ])

        _apply_header_style(ws)
        _apply_detail_header_style(ws, detail_header_row)
        ws.freeze_panes = f"A{detail_header_row + 1}"
        _autofit(ws)

        output_dir = self.config.paths.output_dir / run_context.run_id
        output_dir.mkdir(parents=True, exist_ok=True)
        excel_path = output_dir / f"{result.file_path.stem}.xlsx"
        wb.save(str(excel_path))

        if self.uploader is not None and self.config.onedrive.enabled:
            try:
                document_subtype = fields.get("document_subtype")
                self.uploader.upload_excel(excel_path, document_subtype)
            except Exception as exc:
                import logging
                logging.getLogger("invoice_batch.excel_writer").warning(
                    "No se pudo subir %s a OneDrive: %s", excel_path.name, exc
                )
