import argparse
import json
import os
import sys
from pathlib import Path

from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


load_dotenv()

DEFAULT_INPUT = "factura.pdf"
DEFAULT_EXCEL = "facturas_procesadas.xlsx"
SUPPORTED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".bmp"}


def obtener_valor_seguro(campo):
    return campo.value if campo else None


def obtener_texto(campo):
    valor = obtener_valor_seguro(campo)
    return str(valor) if valor is not None else None


def formatear_direccion(campo):
    valor = obtener_valor_seguro(campo)
    if valor is None:
        return None

    if isinstance(valor, str):
        return valor

    partes = [
        getattr(valor, "street_address", None),
        getattr(valor, "road", None),
        getattr(valor, "house_number", None),
        getattr(valor, "unit", None),
        getattr(valor, "city", None),
        getattr(valor, "state", None),
        getattr(valor, "postal_code", None),
        getattr(valor, "country_region", None),
    ]

    partes_limpias = []
    vistos = set()
    for parte in partes:
        if not parte:
            continue
        parte_normalizada = str(parte).strip()
        clave = parte_normalizada.lower()
        if clave in vistos:
            continue
        vistos.add(clave)
        partes_limpias.append(parte_normalizada)

    if partes_limpias:
        return ", ".join(partes_limpias)

    contenido = getattr(campo, "content", None)
    return contenido.strip() if contenido else str(valor)


def obtener_monto(campo):
    if not campo or not campo.value:
        return None
    return getattr(campo.value, "amount", None)


def obtener_moneda(campo):
    if not campo or not campo.value:
        return None
    codigo = getattr(campo.value, "currency_code", None)
    simbolo = getattr(campo.value, "symbol", None)
    return codigo or simbolo


def obtener_contenido(campo):
    if not campo:
        return None
    return getattr(campo, "content", None)


def crear_cliente():
    endpoint = os.getenv("ENDPOINT")
    key = os.getenv("KEY")

    if not endpoint or not key:
        raise ValueError(
            "Faltan variables de entorno requeridas: ENDPOINT y/o KEY. "
            "Asegurate de cargarlas antes de ejecutar el script."
        )

    return DocumentAnalysisClient(
        endpoint=endpoint,
        credential=AzureKeyCredential(key),
    )


def iterar_archivos_entrada(ruta_entrada):
    ruta = Path(ruta_entrada)

    if not ruta.exists():
        raise FileNotFoundError(f"No existe la ruta indicada: {ruta}")

    if ruta.is_file():
        if ruta.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"Formato no soportado: {ruta.suffix}")
        return [ruta]

    archivos = sorted(
        archivo
        for archivo in ruta.iterdir()
        if archivo.is_file() and archivo.suffix.lower() in SUPPORTED_EXTENSIONS
    )

    if not archivos:
        raise ValueError(f"No se encontraron facturas compatibles en: {ruta}")

    return archivos


def extraer_items(items_field):
    items = []

    if not items_field or not items_field.value:
        return items

    for indice, item in enumerate(items_field.value, start=1):
        item_dict = item.value
        descripcion = item_dict.get("Description")
        cantidad = item_dict.get("Quantity")
        precio_unitario = item_dict.get("UnitPrice")
        importe_linea = item_dict.get("Amount")
        producto_codigo = item_dict.get("ProductCode")
        fecha = item_dict.get("Date")

        items.append(
            {
                "linea": indice,
                "descripcion": obtener_texto(descripcion) or "S/D",
                "cantidad": obtener_valor_seguro(cantidad) or 1,
                "precio_unitario": obtener_monto(precio_unitario) or 0,
                "total_linea": obtener_monto(importe_linea) or 0,
                "codigo_producto": obtener_texto(producto_codigo),
                "fecha_item": obtener_texto(fecha),
            }
        )

    return items


def analizar_factura(client, ruta_archivo):
    with open(ruta_archivo, "rb") as archivo:
        poller = client.begin_analyze_document("prebuilt-invoice", document=archivo)
        result = poller.result()

    if not result.documents:
        raise ValueError("Azure no devolvio documentos analizados para ese archivo.")

    invoice = result.documents[0]
    fields = invoice.fields

    datos_factura = {
        "archivo_origen": Path(ruta_archivo).name,
        "id_factura": obtener_valor_seguro(fields.get("InvoiceId")),
        "orden_compra": obtener_valor_seguro(fields.get("PurchaseOrder")),
        "fecha_emision": obtener_texto(fields.get("InvoiceDate")),
        "fecha_vencimiento": obtener_texto(fields.get("DueDate")),
        "proveedor_nombre": obtener_valor_seguro(fields.get("VendorName")),
        "proveedor_direccion": formatear_direccion(fields.get("VendorAddress")),
        "proveedor_tax_id": obtener_valor_seguro(fields.get("VendorTaxId")),
        "cliente_nombre": obtener_valor_seguro(fields.get("CustomerName")),
        "cliente_id": obtener_valor_seguro(fields.get("CustomerId")),
        "cliente_direccion": formatear_direccion(fields.get("CustomerAddress")),
        "subtotal": obtener_monto(fields.get("SubTotal")),
        "impuesto": obtener_monto(fields.get("TotalTax")),
        "total": obtener_monto(fields.get("InvoiceTotal")),
        "moneda": obtener_moneda(fields.get("InvoiceTotal")),
        "contenido_total": obtener_contenido(fields.get("InvoiceTotal")),
        "items": extraer_items(fields.get("Items")),
    }

    return datos_factura


def guardar_json(datos_factura, carpeta_salida):
    carpeta = Path(carpeta_salida)
    carpeta.mkdir(parents=True, exist_ok=True)

    for factura in datos_factura:
        nombre_base = Path(factura["archivo_origen"]).stem
        salida = carpeta / f"{nombre_base}.json"
        with open(salida, "w", encoding="utf-8") as json_file:
            json.dump(factura, json_file, indent=4, ensure_ascii=False)


def autofit_worksheet(hoja):
    for columna in hoja.columns:
        longitud_maxima = 0
        letra = get_column_letter(columna[0].column)
        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            longitud_maxima = max(longitud_maxima, len(valor))
        hoja.column_dimensions[letra].width = min(longitud_maxima + 2, 40)


def aplicar_estilo_tabla(hoja):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for celda in hoja[1]:
        celda.fill = fill
        celda.font = font

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    autofit_worksheet(hoja)


def exportar_excel(facturas, errores, ruta_excel):
    workbook = Workbook()
    hoja_facturas = workbook.active
    hoja_facturas.title = "Facturas"

    columnas_facturas = [
        "archivo_origen",
        "id_factura",
        "orden_compra",
        "fecha_emision",
        "fecha_vencimiento",
        "proveedor_nombre",
        "proveedor_direccion",
        "proveedor_tax_id",
        "cliente_nombre",
        "cliente_id",
        "cliente_direccion",
        "subtotal",
        "impuesto",
        "total",
        "moneda",
        "contenido_total",
        "cantidad_items",
    ]
    hoja_facturas.append(columnas_facturas)

    for factura in facturas:
        hoja_facturas.append(
            [
                factura.get("archivo_origen"),
                factura.get("id_factura"),
                factura.get("orden_compra"),
                factura.get("fecha_emision"),
                factura.get("fecha_vencimiento"),
                factura.get("proveedor_nombre"),
                factura.get("proveedor_direccion"),
                factura.get("proveedor_tax_id"),
                factura.get("cliente_nombre"),
                factura.get("cliente_id"),
                factura.get("cliente_direccion"),
                factura.get("subtotal"),
                factura.get("impuesto"),
                factura.get("total"),
                factura.get("moneda"),
                factura.get("contenido_total"),
                len(factura.get("items", [])),
            ]
        )

    hoja_items = workbook.create_sheet("Items")
    columnas_items = [
        "archivo_origen",
        "id_factura",
        "linea",
        "descripcion",
        "cantidad",
        "precio_unitario",
        "total_linea",
        "codigo_producto",
        "fecha_item",
    ]
    hoja_items.append(columnas_items)

    for factura in facturas:
        for item in factura.get("items", []):
            hoja_items.append(
                [
                    factura.get("archivo_origen"),
                    factura.get("id_factura"),
                    item.get("linea"),
                    item.get("descripcion"),
                    item.get("cantidad"),
                    item.get("precio_unitario"),
                    item.get("total_linea"),
                    item.get("codigo_producto"),
                    item.get("fecha_item"),
                ]
            )

    hoja_errores = workbook.create_sheet("Errores")
    hoja_errores.append(["archivo_origen", "error"])
    for error in errores:
        hoja_errores.append([error["archivo_origen"], error["error"]])

    for hoja in workbook.worksheets:
        aplicar_estilo_tabla(hoja)

    workbook.save(ruta_excel)


def procesar_facturas(ruta_entrada, ruta_excel, carpeta_json=None):
    client = crear_cliente()
    archivos = iterar_archivos_entrada(ruta_entrada)
    facturas = []
    errores = []

    for archivo in archivos:
        print(f"Procesando: {archivo.name}")
        try:
            factura = analizar_factura(client, archivo)
            facturas.append(factura)
        except Exception as exc:
            errores.append(
                {
                    "archivo_origen": archivo.name,
                    "error": str(exc),
                }
            )
            print(f"  Error: {exc}")

    if carpeta_json:
        guardar_json(facturas, carpeta_json)

    exportar_excel(facturas, errores, ruta_excel)

    return facturas, errores


def parse_args():
    parser = argparse.ArgumentParser(
        description="Lee facturas con Azure Document Intelligence y las exporta a Excel."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=DEFAULT_INPUT,
        help="Ruta a un archivo de factura o a una carpeta con facturas.",
    )
    parser.add_argument(
        "--excel",
        default=DEFAULT_EXCEL,
        help="Nombre o ruta del archivo Excel de salida.",
    )
    parser.add_argument(
        "--json-dir",
        help="Carpeta opcional para guardar un JSON por factura procesada.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    try:
        facturas, errores = procesar_facturas(
            ruta_entrada=args.input,
            ruta_excel=args.excel,
            carpeta_json=args.json_dir,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print("\nResumen")
    print(f"  Facturas procesadas correctamente: {len(facturas)}")
    print(f"  Facturas con error: {len(errores)}")
    print(f"  Excel generado: {args.excel}")

    if args.json_dir:
        print(f"  JSONs generados en: {args.json_dir}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
